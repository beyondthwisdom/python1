import openpyxl

# Excel dosyasını yükle
envanter_dosyasi = openpyxl.load_workbook("inventory.xlsx")
urun_listesi = envanter_dosyasi["Sheet1"]

tedarikci_basina_urun = {}
tedarikci_basina_toplam_deger = {}
onun_altinda_stok_urunler = {}

for urun_satiri in range(2, urun_listesi.max_row + 1):
    tedarikci_adi = urun_listesi.cell(urun_satiri, 4).value
    envanter = urun_listesi.cell(urun_satiri, 2).value
    fiyat = urun_listesi.cell(urun_satiri, 3).value
    urun_numarasi = urun_listesi.cell(urun_satiri, 1).value
    envanter_fiyati = urun_listesi.cell(urun_satiri, 5)

    #print(f"Tedarikçi adı: {tedarikci_adi} envanter: {envanter} fiyat: {fiyat} ürün numarası: {urun_numarasi}")

    # Tedarikçi başına ürün sayısını hesaplama
    if tedarikci_adi in tedarikci_basina_urun:
        mevcut_urun_sayisi = tedarikci_basina_urun.get(tedarikci_adi)
        tedarikci_basina_urun[tedarikci_adi] = mevcut_urun_sayisi + 1
    else:
        tedarikci_basina_urun[tedarikci_adi] = 1

    # Tedarikçi başına toplam envanter değerini hesaplama
    if tedarikci_adi in tedarikci_basina_toplam_deger:
        mevcut_toplam_deger = tedarikci_basina_toplam_deger.get(tedarikci_adi)
        tedarikci_basina_toplam_deger[tedarikci_adi] = mevcut_toplam_deger + envanter * fiyat
    else:
        tedarikci_basina_toplam_deger[tedarikci_adi] = envanter * fiyat

    if envanter < 10:
        onun_altinda_stok_urunler[int(urun_numarasi)] = int(envanter)
    
    # Toplam envanter fiyatını ekleme
    envanter_fiyati.value = envanter * fiyat       

print(tedarikci_basina_urun)
print(tedarikci_basina_toplam_deger)
print(onun_altinda_stok_urunler)    

envanter_dosyasi.save("inventory_with_total_value.xlsx")